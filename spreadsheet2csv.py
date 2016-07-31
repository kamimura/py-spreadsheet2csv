#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import openpyxl
import csv

path = '.'
if len(sys.argv) > 1:
    path = sys.argv[1]

for excel_file in os.listdir(path):
    if (not os.path.isfile(excel_file)) or not excel_file.endswith('.xlsx'):
        continue
    print('{0}: Writing to csv...'.format(excel_file))
    wb = openpyxl.load_workbook(excel_file)
    excel_filename = excel_file[:-5]
    for sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_name)
        csv_filename = '{0}_{1}.csv'.format(excel_filename, sheet_name)
        with open(csv_filename, 'w', newline='') as f:
            writer = csv.writer(f)
            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(
                        sheet.cell(row=row_num, column=col_num).value
                    )
                writer.writerow(row_data)
